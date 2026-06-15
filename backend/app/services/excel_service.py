import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from app.models.user import User, UserRole
from app.core.config import settings

def format_date_with_weekday(dt: datetime) -> str:
    """Форматирует дату в формат: день_недели. день.месяц.год (например: пн. 15.06.2026)"""
    weekdays = ["пн.", "вт.", "ср.", "чт.", "пт.", "сб.", "вс."]
    wd = weekdays[dt.weekday()]
    return f"{wd} {dt.strftime('%d.%m.%Y')}"

def get_local_date(dt_val) -> datetime | None:
    """Извлекает локальную дату из datetime или ISO строки."""
    if pd.isna(dt_val) or not dt_val:
        return None
    if isinstance(dt_val, str):
        try:
            dt = datetime.fromisoformat(dt_val.replace('Z', '+00:00'))
        except ValueError:
            try:
                dt = datetime.strptime(dt_val, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                return None
    else:
        dt = dt_val
        
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(settings.tz_info).date()

async def generate_excel_file(
    data: list,
    current_user: User,
    is_personal: bool = False,
    start_date = None,
    end_date = None
) -> io.BytesIO:
    """Генерирует Excel-файл и возвращает его как BytesIO объект."""
    if not data:
        return None

    # Формируем строку периода для вывода на листах отчета
    period_str = ""
    if start_date or end_date:
        start_fmt = format_date_with_weekday(start_date) if start_date else None
        end_fmt = format_date_with_weekday(end_date) if end_date else None
        
        if start_fmt and end_fmt:
            if start_date == end_date:
                period_str = f"Период: {start_fmt}"
            else:
                period_str = f"Период: {start_fmt} — {end_fmt}"
        elif start_fmt:
            period_str = f"Период: с {start_fmt}"
        elif end_fmt:
            period_str = f"Период: по {end_fmt}"
    else:
        local_now = datetime.now(settings.tz_info)
        period_str = f"Дата: {format_date_with_weekday(local_now)}"

    df = pd.DataFrame(data)
    
    # Колонки и их порядок
    cols_order = ["id", "employee", "project", "start_time", "end_time", "hours", "approved_hours", "description", "status"]
    # Проверяем наличие всех колонок в data, если нет - добавляем пустые
    for col in cols_order:
        if col not in df.columns:
            df[col] = None

    df = df[cols_order].copy()
    
    # Форматирование дат к локальному часовому поясу (Asia/Almaty) с безопасной обработкой None
    def format_datetime_local(dt):
        if pd.isna(dt) or dt is None:
            return ""
        if isinstance(dt, str):
            try:
                dt = pd.to_datetime(dt)
            except Exception:
                return dt
        if hasattr(dt, "to_pydatetime"):
            dt = dt.to_pydatetime()
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(settings.tz_info).strftime('%d.%m.%Y %H:%M')

    df.loc[:, 'start_time'] = df['start_time'].apply(format_datetime_local)
    df.loc[:, 'end_time'] = df['end_time'].apply(format_datetime_local)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for index, row in df.iterrows():
            # Логика заполнения одобренных часов
            if row['status'] == "Подтверждено":
                if pd.isna(row['approved_hours']) or row['approved_hours'] is None:
                    df.at[index, 'approved_hours'] = row['hours']
            else:
                df.at[index, 'approved_hours'] = 0.0
            
            if pd.isna(row['hours']) or row['hours'] is None:
                df.at[index, 'hours'] = 0.0

        # Переименование колонок
        df_export = df.rename(columns={
            "id": "ID",
            "employee": "Сотрудник",
            "project": "Проект",
            "start_time": "Начало",
            "end_time": "Окончание",
            "hours": "Запрошено",
            "approved_hours": "Согласовано",
            "description": "Описание",
            "status": "Статус"
        })

        df_export.to_excel(writer, index=False, sheet_name='Report', startrow=3)
        worksheet = writer.sheets['Report']
        
        # Заголовок
        title_text = "ПЕРСОНАЛЬНЫЙ ОТЧЕТ" if is_personal else "ОТЧЕТ ПО ПЕРЕРАБОТКАМ"
        title = f"{title_text} — {period_str}"
        worksheet.merge_cells('A1:I1')
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True, color="1e40af")
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        worksheet.merge_cells('A2:I2')
        worksheet['A2'] = f"Выгрузил: {current_user.full_name}"
        worksheet['A2'].alignment = Alignment(horizontal='center')

        # Стилизация
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col in range(1, 10):
            cell = worksheet.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_idx in range(5, 5 + len(df)):
            for col_idx in range(1, 10):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx in [6, 7]:
                    cell.alignment = Alignment(horizontal='center')

        # Ширина колонок
        for i, col_name in enumerate(df_export.columns):
            column_letter = get_column_letter(i + 1)
            column_data = df_export[col_name].astype(str)
            max_len = max(column_data.map(len).max(), len(str(col_name))) + 4
            worksheet.column_dimensions[column_letter].width = min(max_len, 50)

        # Итого
        total_row = 5 + len(df)
        worksheet.cell(row=total_row, column=5).value = "ИТОГО:"
        worksheet.cell(row=total_row, column=5).font = Font(bold=True)
        worksheet.cell(row=total_row, column=5).border = border
        
        worksheet.cell(row=total_row, column=6).value = df['hours'].sum()
        worksheet.cell(row=total_row, column=6).font = Font(bold=True)
        worksheet.cell(row=total_row, column=6).border = border
        
        worksheet.cell(row=total_row, column=7).value = df['approved_hours'].sum()
        worksheet.cell(row=total_row, column=7).font = Font(bold=True, color="15803d")
        worksheet.cell(row=total_row, column=7).border = border
        
        for col_idx in [1, 2, 3, 4, 8, 9]:
            worksheet.cell(row=total_row, column=col_idx).border = border

        # Группировка сотрудников для дополнительных листов (табелей)
        if not is_personal and current_user.role in [UserRole.manager, UserRole.head, UserRole.admin]:
            # Шаг 1: Посчитаем суммарные согласованные часы каждого сотрудника для определения листа (>16 или <=16)
            emp_totals = {}
            for item in data:
                emp_name = item.get("employee")
                if not emp_name:
                    continue
                if emp_name not in emp_totals:
                    emp_totals[emp_name] = {
                        "company": item.get("employee_company") or "Polymedia",
                        "total_approved": 0.0,
                        "department": item.get("department") or ""
                    }
                if item.get("status") == "Подтверждено":
                    approved = item.get("approved_hours")
                    if approved is None or pd.isna(approved):
                        approved = item.get("hours", 0.0) or 0.0
                    emp_totals[emp_name]["total_approved"] += approved

            # Шаг 2: Распределим сотрудников по 4 группам на основе суммарных часов и компании
            polymedia_gt16_members = set()
            polymedia_lte16_members = set()
            ajtech_gt16_members = set()
            ajtech_lte16_members = set()
            
            for emp_name, info in emp_totals.items():
                total_approved = info["total_approved"]
                company = info["company"]
                if total_approved > 16:
                    if company == "Polymedia":
                        polymedia_gt16_members.add(emp_name)
                    else:
                        ajtech_gt16_members.add(emp_name)
                else:
                    if company == "Polymedia":
                        polymedia_lte16_members.add(emp_name)
                    else:
                        ajtech_lte16_members.add(emp_name)

            # Вспомогательные функции для подготовки данных
            def get_group_dates(members):
                group_dates = set()
                for item in data:
                    emp_name = item.get("employee")
                    if emp_name in members and item.get("status") == "Подтверждено":
                        emp_date = get_local_date(item.get("start_time"))
                        if emp_date:
                            group_dates.add(emp_date)
                return sorted(list(group_dates))

            def prepare_group_data(members):
                # Нам нужны только те сотрудники, у которых есть согласованные часы > 0
                active_members = sorted([m for m in members if m in emp_totals and emp_totals[m]["total_approved"] > 0])
                group_dates = get_group_dates(active_members)
                
                structured_data = []
                for emp_name in active_members:
                    emp_info = {
                        "employee": emp_name,
                        "department": emp_totals[emp_name]["department"],
                        "projects": {}
                    }
                    
                    for item in data:
                        if item.get("employee") == emp_name and item.get("status") == "Подтверждено":
                            proj = item.get("project") or "Внутренний"
                            emp_date = get_local_date(item.get("start_time"))
                            if not emp_date:
                                continue
                            
                            approved = item.get("approved_hours")
                            if approved is None or pd.isna(approved):
                                approved = item.get("hours", 0.0) or 0.0
                                
                            if proj not in emp_info["projects"]:
                                emp_info["projects"][proj] = {d: 0.0 for d in group_dates}
                            
                            emp_info["projects"][proj][emp_date] += approved
                    
                    structured_data.append(emp_info)
                
                return group_dates, structured_data

            # Функция для генерации и форматирования листа табеля с датами по горизонтали
            def create_timesheet_sheet(sheet_name, title_desc, members):
                group_dates, structured_data = prepare_group_data(members)
                ws = writer.book.create_sheet(title=sheet_name)
                
                # Заголовок листа
                headers = ["№", "Сотрудник", "Отдел", "Проекты"]
                for d in group_dates:
                    headers.append(format_date_with_weekday(d))
                headers.append("Итого")
                
                num_cols = len(headers)
                last_col_letter = get_column_letter(num_cols)
                
                ws.merge_cells(f'A1:{last_col_letter}1')
                ws['A1'] = f"ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ — {title_desc}"
                ws['A1'].font = Font(size=14, bold=True, color="1e40af")
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 30
                
                ws.merge_cells(f'A2:{last_col_letter}2')
                ws['A2'] = f"Выгрузил: {current_user.full_name} | {period_str}"
                ws['A2'].font = Font(italic=True, size=10, color="4b5563")
                ws['A2'].alignment = Alignment(horizontal='center')
                ws.row_dimensions[2].height = 20

                # Если нет активных сотрудников или дат, выводим сообщение
                if not structured_data or not group_dates:
                    ws.merge_cells(f'A4:{last_col_letter}4')
                    ws['A4'] = "Нет данных за выбранный период"
                    ws['A4'].font = Font(italic=True, size=11, color="4b5563")
                    ws['A4'].alignment = Alignment(horizontal='center')
                    return

                # Заголовки таблицы
                header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                ws.row_dimensions[4].height = 25
                for col_idx, h_text in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_idx, value=h_text)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border_thin

                # Данные
                row_idx = 5
                for emp_idx, emp_info in enumerate(structured_data, 1):
                    projects = sorted(list(emp_info["projects"].keys()))
                    if not projects:
                        continue
                    
                    emp_daily_totals = {d: 0.0 for d in group_dates}
                    emp_grand_total = 0.0
                    
                    # Выводим строки по проектам
                    for proj_idx, proj in enumerate(projects):
                        ws.row_dimensions[row_idx].height = 20
                        
                        if proj_idx == 0:
                            ws.cell(row=row_idx, column=1, value=emp_idx).alignment = Alignment(horizontal='center')
                            ws.cell(row=row_idx, column=2, value=emp_info["employee"])
                            ws.cell(row=row_idx, column=3, value=emp_info["department"])
                        else:
                            ws.cell(row=row_idx, column=1, value="")
                            ws.cell(row=row_idx, column=2, value="")
                            ws.cell(row=row_idx, column=3, value="")
                            
                        ws.cell(row=row_idx, column=4, value=proj)
                        
                        proj_total = 0.0
                        for d_idx, d in enumerate(group_dates, 5):
                            val = emp_info["projects"][proj].get(d, 0.0)
                            emp_daily_totals[d] += val
                            proj_total += val
                            cell_val = val if val > 0 else ""
                            ws.cell(row=row_idx, column=d_idx, value=cell_val).alignment = Alignment(horizontal='center')
                            
                        emp_grand_total += proj_total
                        proj_total_val = proj_total if proj_total > 0 else ""
                        ws.cell(row=row_idx, column=num_cols, value=proj_total_val).font = Font(bold=True)
                        ws.cell(row=row_idx, column=num_cols).alignment = Alignment(horizontal='center')
                        
                        for col_idx in range(1, num_cols + 1):
                            ws.cell(row=row_idx, column=col_idx).border = border_thin
                            
                        row_idx += 1
                        
                    # Выводим строку "Итого по сотруднику"
                    ws.row_dimensions[row_idx].height = 22
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="Итого по сотруднику:").font = Font(bold=True)
                    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(row=row_idx, column=3, value="")
                    ws.cell(row=row_idx, column=4, value="")
                    
                    for d_idx, d in enumerate(group_dates, 5):
                        val = emp_daily_totals[d]
                        cell_val = val if val > 0 else ""
                        ws.cell(row=row_idx, column=d_idx, value=cell_val).font = Font(bold=True)
                        ws.cell(row=row_idx, column=d_idx).alignment = Alignment(horizontal='center')
                        
                    ws.cell(row=row_idx, column=num_cols, value=emp_grand_total).font = Font(bold=True)
                    ws.cell(row=row_idx, column=num_cols).alignment = Alignment(horizontal='center')
                    
                    subtotal_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
                    for col_idx in range(1, num_cols + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = border_thin
                        cell.fill = subtotal_fill
                        
                    row_idx += 1

                # Общая итоговая строка по всему листу
                ws.row_dimensions[row_idx].height = 24
                ws.cell(row=row_idx, column=4, value="ОБЩИЙ ИТОГО:").font = Font(bold=True)
                ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='right', vertical='center')
                
                sheet_daily_totals = {d: 0.0 for d in group_dates}
                sheet_grand_total = 0.0
                for emp_info in structured_data:
                    for proj, dates_dict in emp_info["projects"].items():
                        for d, val in dates_dict.items():
                            sheet_daily_totals[d] += val
                            sheet_grand_total += val
                            
                for d_idx, d in enumerate(group_dates, 5):
                    val = sheet_daily_totals[d]
                    cell_val = val if val > 0 else ""
                    ws.cell(row=row_idx, column=d_idx, value=cell_val).font = Font(bold=True)
                    ws.cell(row=row_idx, column=d_idx).alignment = Alignment(horizontal='center')
                    
                ws.cell(row=row_idx, column=num_cols, value=sheet_grand_total).font = Font(bold=True, color="15803d")
                ws.cell(row=row_idx, column=num_cols).alignment = Alignment(horizontal='center')
                
                for col_idx in range(1, num_cols + 1):
                    ws.cell(row=row_idx, column=col_idx).border = border_thin
                
                # Установка ширин колонок
                ws.column_dimensions["A"].width = 6
                ws.column_dimensions["B"].width = 30
                ws.column_dimensions["C"].width = 25
                ws.column_dimensions["D"].width = 25
                for col_idx in range(5, num_cols):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 14
                ws.column_dimensions[get_column_letter(num_cols)].width = 12

            # Создаем 4 дополнительных листа
            create_timesheet_sheet("Polymedia (>16ч)", "Polymedia (более 16ч)", polymedia_gt16_members)
            create_timesheet_sheet("Polymedia (<=16ч)", "Polymedia (до 16ч)", polymedia_lte16_members)
            create_timesheet_sheet("AJ-techCom (>16ч)", "AJ-techCom (более 16ч)", ajtech_gt16_members)
            create_timesheet_sheet("AJ-techCom (<=16ч)", "AJ-techCom (до 16ч)", ajtech_lte16_members)

    output.seek(0)
    return output
