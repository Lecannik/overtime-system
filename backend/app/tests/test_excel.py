import pytest
from datetime import datetime, timezone
import openpyxl
import io
from app.services.excel_service import format_date_with_weekday, generate_excel_file
from app.models.user import User, UserRole, UserCompany
from app.models.organization import Department

def test_format_date_with_weekday():
    # 12 июня 2026 года - пятница (пт.)
    dt_fri = datetime(2026, 6, 12)
    assert format_date_with_weekday(dt_fri) == "пт. 12.06.2026"

    # 15 июня 2026 года - понедельник (пн.)
    dt_mon = datetime(2026, 6, 15)
    assert format_date_with_weekday(dt_mon) == "пн. 15.06.2026"

    # 14 июня 2026 года - воскресенье (вс.)
    dt_sun = datetime(2026, 6, 14)
    assert format_date_with_weekday(dt_sun) == "вс. 14.06.2026"

@pytest.mark.asyncio
async def test_generate_excel_file():
    # Подготовим фейковые данные для экспорта
    fake_user = User(
        id=1,
        full_name="Иван Иванов",
        email="admin@example.com",
        role=UserRole.admin,
        company=UserCompany.Polymedia,
        is_active=True
    )
    
    data = [
        {
            "id": 1,
            "employee": "Чмиль Никита Павлович",
            "employee_company": "Polymedia",
            "department": "Технический отдел",
            "project": "Test_Projject",
            "start_time": "2026-06-12 18:00:00",
            "end_time": "2026-06-12 21:00:00",
            "hours": 3.0,
            "approved_hours": 3.0,
            "description": "Тестовая переработка",
            "status": "Подтверждено"
        },
        {
            "id": 2,
            "employee": "Дощанов Руслан Аскарович",
            "employee_company": "Polymedia",
            "department": "Отдел программных решений",
            "project": "Внутренний",
            "start_time": "2026-06-12 18:00:00",
            "end_time": "2026-06-12 20:00:00",
            "hours": 2.0,
            "approved_hours": 2.0,
            "description": "Внутренняя работа",
            "status": "Подтверждено"
        },
        {
            "id": 3,
            "employee": "Петров Петр Петрович",
            "employee_company": "AJ-techCom",
            "department": "Отдел автоматизации",
            "project": "AJ_Project",
            "start_time": "2026-06-12 10:00:00",
            "end_time": "2026-06-12 15:00:00",
            "hours": 5.0,
            "approved_hours": 5.0,
            "description": "Автоматизация процессов",
            "status": "Подтверждено"
        }
    ]

    # 1. Проверяем генерацию БЕЗ указания дат (должна выводиться текущая Дата)
    excel_data = await generate_excel_file(data, fake_user, is_personal=False)
    assert excel_data is not None
    assert isinstance(excel_data, io.BytesIO)

    wb = openpyxl.load_workbook(excel_data)
    ws_poly = wb["Polymedia (<=16ч)"]
    a2_val = ws_poly['A2'].value
    assert "Выгрузил: Иван Иванов" in a2_val
    assert "Дата: " in a2_val

    # Проверим, что создался лист для AJ-techCom
    ws_aj_check = wb["AJ-techCom (<=16ч)"]
    assert ws_aj_check is not None
    assert ws_aj_check.cell(row=5, column=2).value == "Петров Петр Петрович"
    assert ws_aj_check.cell(row=5, column=5).value == 5.0

    # 2. Проверяем генерацию С указанием периода (должен выводиться Период)
    from datetime import date
    start_d = date(2026, 6, 1)
    end_d = date(2026, 6, 12)
    excel_data_period = await generate_excel_file(
        data, fake_user, is_personal=False, start_date=start_d, end_date=end_d
    )
    
    wb_period = openpyxl.load_workbook(excel_data_period)
    ws_poly_period = wb_period["Polymedia (<=16ч)"]
    a2_period_val = ws_poly_period['A2'].value
    
    # Период: пн. 01.06.2026 — пт. 12.06.2026
    assert "Выгрузил: Иван Иванов" in a2_period_val
    assert "Период: пн. 01.06.2026 — пт. 12.06.2026" in a2_period_val

    # Проверим заголовки таблицы (в строке 4)
    # headers = ["№", "Сотрудник", "Отдел", "Проекты", "пт. 12.06.2026", "Итого"]
    headers = [ws_poly_period.cell(row=4, column=col).value for col in range(1, 7)]
    assert headers == ["№", "Сотрудник", "Отдел", "Проекты", "пт. 12.06.2026", "Итого"]

    # Проверим первую строку данных (строка 5) для первого сотрудника по алфавиту (Дощанов Руслан Аскарович)
    assert ws_poly_period.cell(row=5, column=1).value == 1
    assert ws_poly_period.cell(row=5, column=2).value == "Дощанов Руслан Аскарович"
    assert ws_poly_period.cell(row=5, column=3).value == "Отдел программных решений"
    assert ws_poly_period.cell(row=5, column=4).value == "Внутренний"
    assert ws_poly_period.cell(row=5, column=5).value == 2.0
    assert ws_poly_period.cell(row=5, column=6).value == 2.0

    # Строка 6 - Итого по первому сотруднику
    assert ws_poly_period.cell(row=6, column=2).value == "Итого по сотруднику:"
    assert ws_poly_period.cell(row=6, column=5).value == 2.0
    assert ws_poly_period.cell(row=6, column=6).value == 2.0

    # Строка 7 - второй сотрудник (Чмиль Никита Павлович)
    assert ws_poly_period.cell(row=7, column=1).value == 2
    assert ws_poly_period.cell(row=7, column=2).value == "Чмиль Никита Павлович"
    assert ws_poly_period.cell(row=7, column=3).value == "Технический отдел"
    assert ws_poly_period.cell(row=7, column=4).value == "Test_Projject"
    assert ws_poly_period.cell(row=7, column=5).value == 3.0
    assert ws_poly_period.cell(row=7, column=6).value == 3.0

    # Строка 8 - Итого по второму сотруднику
    assert ws_poly_period.cell(row=8, column=2).value == "Итого по сотруднику:"
    assert ws_poly_period.cell(row=8, column=5).value == 3.0
    assert ws_poly_period.cell(row=8, column=6).value == 3.0

    # Строка 9 - ОБЩИЙ ИТОГО
    assert ws_poly_period.cell(row=9, column=4).value == "ОБЩИЙ ИТОГО:"
    assert ws_poly_period.cell(row=9, column=5).value == 5.0
    assert ws_poly_period.cell(row=9, column=6).value == 5.0
